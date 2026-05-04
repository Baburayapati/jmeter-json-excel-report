import json
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter


REMOVE_COLUMNS = [
    "medianResTime",
    "throughput",
    "receivedKBytesPerSec",
    "sentKBytesPerSec",
]

TIME_COLUMNS_MS_TO_SEC = {
    "meanResTime": "Avg ResTime in sec",
    "minResTime": "Min ResTime in sec",
    "maxResTime": "MaxRes Time in sec",
    "pct1ResTime": "90thPercentile Resp Time in Sec",
    "pct2ResTime": "95thPercentile Resp Time in Sec",
    "pct3ResTime": "99thPercentile Resp Time in Sec",
}


def is_transaction(name: str) -> bool:
    return bool(re.match(r"^T\d{2}", str(name).strip()))


def split_api_name(name: str) -> Tuple[str, str, str]:
    parts = str(name).split("/")
    if len(parts) >= 3:
        return parts[0], parts[1], "/".join(parts[2:])
    if len(parts) == 2:
        return parts[0], parts[1], ""
    return str(name), "", ""


def load_statistics_json(json_path: str | Path) -> pd.DataFrame:
    with open(json_path, "r", encoding="utf-8") as file:
        data: Dict[str, Dict[str, Any]] = json.load(file)

    rows: List[Dict[str, Any]] = []
    for key, value in data.items():
        row = dict(value)
        row["transaction"] = row.get("transaction", key)
        rows.append(row)

    return pd.DataFrame(rows)


def apply_common_column_cleanup(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    for col in REMOVE_COLUMNS:
        if col in df.columns:
            df = df.drop(columns=[col])

    for old_col, new_col in TIME_COLUMNS_MS_TO_SEC.items():
        if old_col in df.columns:
            df[old_col] = pd.to_numeric(df[old_col], errors="coerce") / 1000
            df = df.rename(columns={old_col: new_col})

    return df


def add_api_sla_columns(apis_df: pd.DataFrame) -> pd.DataFrame:
    apis_df = apis_df.copy()
    apis_df["SLA Sec"] = apis_df["Feature"].astype(str).str.upper().str.startswith("ASKAI").map({True: 10, False: 2})
    apis_df["SLA Rule"] = apis_df["SLA Sec"].map(
        lambda x: "AskAI APIs SLA < 10 sec" if x == 10 else "Assets, Assessments, Home, Settings and Support APIs SLA < 2 sec"
    )
    apis_df["SLA Status"] = apis_df.apply(
        lambda row: "PASS"
        if pd.to_numeric(row.get("Avg ResTime in sec"), errors="coerce") < row["SLA Sec"]
        else "FAIL",
        axis=1,
    )
    apis_df["SLA Breach Sec"] = apis_df.apply(
        lambda row: max((pd.to_numeric(row.get("Avg ResTime in sec"), errors="coerce") or 0) - row["SLA Sec"], 0),
        axis=1,
    )
    return apis_df


def order_columns(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    if sheet_name == "APIs":
        # Do not show SLA helper columns in APIs sheet.
        # SLA calculations are still used for response-time cell highlighting.
        preferred = [
            "Feature",
            "Scenario",
            "Endpoint",
            "sampleCount",
            "errorCount",
            "errorPct",
            "Avg ResTime in sec",
            "Min ResTime in sec",
            "MaxRes Time in sec",
            "90thPercentile Resp Time in Sec",
            "95thPercentile Resp Time in Sec",
            "99thPercentile Resp Time in Sec",
        ]
        hidden = {"transaction", "SLA Sec", "SLA Rule", "SLA Status", "SLA Breach Sec"}
        remaining = [c for c in df.columns if c not in preferred and c not in hidden]
        return df[[c for c in preferred if c in df.columns] + remaining]

    preferred = [
        "transaction",
        "sampleCount",
        "errorCount",
        "errorPct",
        "Avg ResTime in sec",
        "Min ResTime in sec",
        "MaxRes Time in sec",
        "90thPercentile Resp Time in Sec",
        "95thPercentile Resp Time in Sec",
        "99thPercentile Resp Time in Sec",
    ]
    remaining = [c for c in df.columns if c not in preferred]
    return df[[c for c in preferred if c in df.columns] + remaining]


def build_single_report_frames(json_path: str | Path):
    df = load_statistics_json(json_path)

    transactions_df = df[df["transaction"].apply(is_transaction)].copy()
    apis_df = df[~df["transaction"].apply(is_transaction)].copy()
    errors_df = df[pd.to_numeric(df.get("errorCount", 0), errors="coerce").fillna(0) > 0].copy()

    split_df = apis_df["transaction"].apply(lambda x: pd.Series(split_api_name(x)))
    split_df.columns = ["Feature", "Scenario", "Endpoint"]
    apis_df = pd.concat([split_df, apis_df], axis=1)

    transactions_df = order_columns(apply_common_column_cleanup(transactions_df), "Transactions")
    errors_df = order_columns(apply_common_column_cleanup(errors_df), "Errors")
    apis_df = order_columns(add_api_sla_columns(apply_common_column_cleanup(apis_df)), "APIs")

    return {
        "Transactions": transactions_df,
        "Errors": errors_df,
        "APIs": apis_df,
    }


def bucket_headers_for_track(track: str) -> List[str]:
    if str(track).upper().startswith("ASKAI"):
        return ["0 - 10s in %", "10 - 20s in %", "20 - 30s in %", "> 30s in %"]
    return ["0 - 2s in %", "3 - 4s in %", "4 - 6s in %", "> 6s in %"]


def bucket_index(seconds: float, is_askai: bool) -> int:
    value = float(seconds)
    if is_askai:
        if value <= 10:
            return 0
        if value <= 20:
            return 1
        if value <= 30:
            return 2
        return 3

    if value <= 2:
        return 0
    if value <= 4:
        return 1
    if value <= 6:
        return 2
    return 3


def prepare_api_df_for_track(json_path: str | Path) -> pd.DataFrame:
    df = load_statistics_json(json_path)
    df = df[~df["transaction"].apply(is_transaction)].copy()
    split_df = df["transaction"].apply(lambda x: pd.Series(split_api_name(x)))
    split_df.columns = ["Feature", "Scenario", "Endpoint"]
    df = pd.concat([split_df, df], axis=1)

    df["avg_sec"] = pd.to_numeric(df["meanResTime"], errors="coerce") / 1000
    df["min_sec"] = pd.to_numeric(df["minResTime"], errors="coerce") / 1000
    df["max_sec"] = pd.to_numeric(df["maxResTime"], errors="coerce") / 1000
    return df


def track_metric_values(df: pd.DataFrame, track: str, metric: str) -> List[Any]:
    g = df[df["Feature"] == track].copy()
    if g.empty:
        return ["", "", "", "", ""]

    metric_to_col = {
        "Avg": "avg_sec",
        "Min": "min_sec",
        "Max": "max_sec",
    }
    col = metric_to_col[metric]
    is_askai = str(track).upper().startswith("ASKAI")

    total_apis = len(g)
    counts = [0, 0, 0, 0]

    for value in pd.to_numeric(g[col], errors="coerce").dropna():
        counts[bucket_index(value, is_askai)] += 1

    percentages = [round((count / total_apis) * 100, 2) for count in counts]
    max_seconds = round(float(pd.to_numeric(g["max_sec"], errors="coerce").max()), 2)
    return percentages + [max_seconds]


def build_track_comparison_matrix(json_paths: List[str | Path], labels: List[str]) -> List[List[Any]]:
    prepared = [prepare_api_df_for_track(path) for path in json_paths]
    all_tracks = sorted(
        track
        for track in set().union(*[set(df["Feature"].dropna().astype(str)) for df in prepared])
        if track
        and track.strip().lower() != "total"
        and "select customer" not in track.strip().lower()
    )

    # 2-row header: run label row, then bucket row.
    header_1 = ["Track", "Metric"]
    header_2 = ["Track", "Metric"]

    for label in labels:
        header_1 += [label, "", "", "", "", ""]
        header_2 += [
            "0-10s AskAI / 0-2s Other %",
            "10-20s AskAI / 3-4s Other %",
            "20-30s AskAI / 4-6s Other %",
            ">30s AskAI / >6s Other %",
            "Max Seconds",
            "",
        ]

    matrix = [header_1, header_2]

    for track in all_tracks:
        for metric in ["Avg", "Min", "Max"]:
            row = [track if metric == "Avg" else "", metric]
            for df in prepared:
                row += track_metric_values(df, track, metric)
                row += [""]
            matrix.append(row)

    return matrix


def prepare_compare_frame(df: pd.DataFrame, label: str) -> pd.DataFrame:
    base = df.copy()
    base["transaction"] = base["transaction"].astype(str)
    needed = [
        "transaction",
        "sampleCount",
        "errorCount",
        "errorPct",
        "meanResTime",
        "pct1ResTime",
        "pct2ResTime",
        "pct3ResTime",
    ]
    available = [c for c in needed if c in base.columns]
    base = base[available].copy()

    for col in ["meanResTime", "pct1ResTime", "pct2ResTime", "pct3ResTime"]:
        if col in base.columns:
            base[col] = pd.to_numeric(base[col], errors="coerce") / 1000

    rename_map = {
        "sampleCount": f"{label} Sample Count",
        "errorCount": f"{label} Error Count",
        "errorPct": f"{label} Error %",
        "meanResTime": f"{label} Avg Sec",
        "pct1ResTime": f"{label} 90th Sec",
        "pct2ResTime": f"{label} 95th Sec",
        "pct3ResTime": f"{label} 99th Sec",
    }
    return base.rename(columns=rename_map)


def build_comparison(json_paths: List[str | Path], labels: List[str]) -> pd.DataFrame:
    compare_df = prepare_compare_frame(load_statistics_json(json_paths[0]), labels[0])

    for path, label in zip(json_paths[1:], labels[1:]):
        next_df = prepare_compare_frame(load_statistics_json(path), label)
        compare_df = compare_df.merge(next_df, on="transaction", how="outer")

    baseline = labels[0]
    latest = labels[-1]

    if f"{baseline} Avg Sec" in compare_df.columns and f"{latest} Avg Sec" in compare_df.columns:
        compare_df["Avg Sec Diff"] = compare_df[f"{latest} Avg Sec"] - compare_df[f"{baseline} Avg Sec"]
        compare_df["Avg Sec Diff %"] = (compare_df["Avg Sec Diff"] / compare_df[f"{baseline} Avg Sec"]) * 100

    if f"{baseline} 90th Sec" in compare_df.columns and f"{latest} 90th Sec" in compare_df.columns:
        compare_df["90th Sec Diff"] = compare_df[f"{latest} 90th Sec"] - compare_df[f"{baseline} 90th Sec"]
        compare_df["90th Sec Diff %"] = (compare_df["90th Sec Diff"] / compare_df[f"{baseline} 90th Sec"]) * 100

    if f"{baseline} Error Count" in compare_df.columns and f"{latest} Error Count" in compare_df.columns:
        compare_df["Error Count Diff"] = compare_df[f"{latest} Error Count"] - compare_df[f"{baseline} Error Count"]

    split_df = compare_df["transaction"].apply(lambda x: pd.Series(split_api_name(x)))
    split_df.columns = ["Feature", "Scenario", "Endpoint"]
    compare_df = pd.concat([split_df, compare_df], axis=1)

    first_cols = ["Feature", "Scenario", "Endpoint", "transaction"]
    other_cols = [c for c in compare_df.columns if c not in first_cols]
    return compare_df[first_cols + other_cols]


def build_report(json_path: str | Path, output_excel_path: str | Path) -> None:
    frames = build_single_report_frames(json_path)
    track_matrix = build_track_comparison_matrix([json_path], [Path(json_path).stem])
    write_excel(frames, output_excel_path, track_matrix=track_matrix)



def build_apis_comparison(json_paths: List[str | Path], labels: List[str]) -> pd.DataFrame:
    """
    Build API-level side-by-side comparison.
    Each uploaded report gets its own Feature, Scenario, Endpoint columns
    plus selected metric columns.
    """
    comparison_df = None

    metric_cols = [
        "sampleCount",
        "errorCount",
        "errorPct",
        "meanResTime",
        "minResTime",
        "maxResTime",
        "pct1ResTime",
        "pct2ResTime",
        "pct3ResTime",
    ]

    metric_rename = {
        "sampleCount": "Sample Count",
        "errorCount": "Error Count",
        "errorPct": "Error %",
        "meanResTime": "Avg ResTime in sec",
        "minResTime": "Min ResTime in sec",
        "maxResTime": "MaxRes Time in sec",
        "pct1ResTime": "90thPercentile Resp Time in Sec",
        "pct2ResTime": "95thPercentile Resp Time in Sec",
        "pct3ResTime": "99thPercentile Resp Time in Sec",
    }

    for path, label in zip(json_paths, labels):
        df = load_statistics_json(path)
        df = df[~df["transaction"].apply(is_transaction)].copy()

        split_df = df["transaction"].apply(lambda x: pd.Series(split_api_name(x)))
        split_df.columns = [f"{label} Feature", f"{label} Scenario", f"{label} Endpoint"]
        df = pd.concat([split_df, df], axis=1)

        # Convert response-time metrics from ms to sec.
        for col in ["meanResTime", "minResTime", "maxResTime", "pct1ResTime", "pct2ResTime", "pct3ResTime"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce") / 1000

        keep_cols = ["transaction", f"{label} Feature", f"{label} Scenario", f"{label} Endpoint"]
        keep_cols += [c for c in metric_cols if c in df.columns]
        run_df = df[keep_cols].copy()

        rename_map = {
            old: f"{label} {new}"
            for old, new in metric_rename.items()
            if old in run_df.columns
        }
        run_df = run_df.rename(columns=rename_map)

        if comparison_df is None:
            comparison_df = run_df
        else:
            comparison_df = comparison_df.merge(run_df, on="transaction", how="outer")

    if comparison_df is None:
        return pd.DataFrame()

    # Add diff columns comparing first and last upload.
    baseline = labels[0]
    latest = labels[-1]

    base_avg = f"{baseline} Avg ResTime in sec"
    latest_avg = f"{latest} Avg ResTime in sec"
    if base_avg in comparison_df.columns and latest_avg in comparison_df.columns:
        comparison_df["Avg ResTime Diff Sec"] = comparison_df[latest_avg] - comparison_df[base_avg]
        comparison_df["Avg ResTime Diff %"] = (comparison_df["Avg ResTime Diff Sec"] / comparison_df[base_avg]) * 100

    base_err = f"{baseline} Error Count"
    latest_err = f"{latest} Error Count"
    if base_err in comparison_df.columns and latest_err in comparison_df.columns:
        comparison_df["Error Count Diff"] = comparison_df[latest_err] - comparison_df[base_err]

    # Put transaction at end because Feature/Scenario/Endpoint columns are the main view.
    ordered = [c for c in comparison_df.columns if c != "transaction"] + ["transaction"]
    return comparison_df[ordered]


def build_comparison_report(json_paths: List[str | Path], labels: List[str], output_excel_path: str | Path) -> None:
    # For comparison mode, keep the workbook focused:
    # Insights + Track_Comparison + APIs_Comparison only.
    latest_frames = build_single_report_frames(json_paths[-1])
    frames = {
        "APIs_Comparison": build_apis_comparison(json_paths, labels)
    }
    track_matrix = build_track_comparison_matrix(json_paths, labels)
    write_excel(
        frames,
        output_excel_path,
        track_matrix=track_matrix,
        insights_frames=latest_frames,
        comparison_mode=True,
    )


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(color="FFFFFF", bold=True)
    dark_font = Font(color="000000", bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A3" if ws.title == "Track_Comparison" else "A2"

    max_header_rows = 2 if ws.title == "Track_Comparison" else 1
    for row_idx in range(1, max_header_rows + 1):
        for cell in ws[row_idx]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.fill = subheader_fill
                cell.font = dark_font

    for row in ws.iter_rows(min_row=max_header_rows + 1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.00"

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        width = min(max(max_len + 2, 12), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Highlight only response-time cells that breach SLA in APIs sheet.
    headers = [cell.value for cell in ws[1]]
    if ws.title == "APIs" and "SLA Sec" in headers:
        sla_col = headers.index("SLA Sec") + 1
        sla_status_col = headers.index("SLA Status") + 1 if "SLA Status" in headers else None
        response_time_columns = [
            "Avg ResTime in sec",
            "Min ResTime in sec",
            "MaxRes Time in sec",
            "90thPercentile Resp Time in Sec",
            "95thPercentile Resp Time in Sec",
            "99thPercentile Resp Time in Sec",
        ]
        target_cols = [headers.index(col_name) + 1 for col_name in response_time_columns if col_name in headers]

        for row in range(2, ws.max_row + 1):
            sla_value = ws.cell(row=row, column=sla_col).value
            try:
                sla_value = float(sla_value)
            except (TypeError, ValueError):
                continue

            for col in target_cols:
                cell = ws.cell(row=row, column=col)
                try:
                    metric_value = float(cell.value)
                except (TypeError, ValueError):
                    continue

                if metric_value >= sla_value:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(color="9C0006", bold=True)

            if sla_status_col:
                status_cell = ws.cell(row=row, column=sla_status_col)
                if status_cell.value == "PASS":
                    status_cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    status_cell.font = Font(color="006100", bold=True)
                elif status_cell.value == "FAIL":
                    status_cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    status_cell.font = Font(color="9C0006", bold=True)


    # APIs: highlight ONLY response-time metric cells that breach SLA.
    # AskAI Feature => 10 sec SLA. Other tracks => 2 sec SLA.
    if ws.title == "APIs":
        headers = [cell.value for cell in ws[1]]
        metric_cols = []
        for col_name in [
            "Avg ResTime in sec",
            "Min ResTime in sec",
            "MaxRes Time in sec",
            "90thPercentile Resp Time in Sec",
            "95thPercentile Resp Time in Sec",
            "99thPercentile Resp Time in Sec",
        ]:
            if col_name in headers:
                metric_cols.append(headers.index(col_name) + 1)

        feature_col = headers.index("Feature") + 1 if "Feature" in headers else None

        if feature_col and metric_cols:
            for row in range(2, ws.max_row + 1):
                feature = str(ws.cell(row=row, column=feature_col).value or "")
                sla_sec = 10 if feature.upper().startswith("ASKAI") else 2

                for col in metric_cols:
                    cell = ws.cell(row=row, column=col)
                    try:
                        value = float(cell.value)
                    except Exception:
                        continue
                    if value >= sla_sec:
                        cell.fill = PatternFill("solid", fgColor="FFC7CE")
                        cell.font = Font(color="9C0006", bold=True)


    # APIs_Comparison: format response-time, error, and diff columns.
    if ws.title == "APIs_Comparison":
        headers = [cell.value for cell in ws[1]]
        diff_cols = []
        for idx, header in enumerate(headers, start=1):
            h = str(header or "").lower()
            if "sec" in h or "error %" in h or "diff %" in h:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=idx).number_format = "0.00"
            if "diff" in h:
                diff_cols.append(idx)

        for col in diff_cols:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    value = float(cell.value)
                except Exception:
                    continue
                if value > 0:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(color="9C0006", bold=True)
                elif value < 0:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(color="006100", bold=True)

    # Highlight Max Seconds columns in Track_Comparison.
    if ws.title == "Track_Comparison":
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=2, column=col).value == "Max Seconds":
                for row in range(3, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F4B6C2")
                    ws.cell(row=row, column=col).font = Font(color="9C0006", bold=True)


def build_insights_sheet(ws, frames: Dict[str, pd.DataFrame]):
    ws.title = "Insights"
    ws["A1"] = "JMeter Performance Insights"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws.merge_cells("A1:F1")

    apis_df = frames["APIs"]
    errors_df = frames["Errors"]
    tx_df = frames["Transactions"]

    total_apis = len(apis_df)
    total_tx = len(tx_df)
    total_error_count = (
        int(pd.to_numeric(errors_df.get("errorCount", 0), errors="coerce").fillna(0).sum()) if not errors_df.empty else 0
    )
    if not apis_df.empty:
        sla_sec_series = apis_df["Feature"].astype(str).str.upper().str.startswith("ASKAI").map({True: 10, False: 2})
        avg_sec_series = pd.to_numeric(apis_df.get("Avg ResTime in sec", 0), errors="coerce").fillna(0)
        sla_pass = int((avg_sec_series < sla_sec_series).sum())
        sla_fail = int((avg_sec_series >= sla_sec_series).sum())
    else:
        sla_pass = 0
        sla_fail = 0
    avg_resp = (
        round(float(pd.to_numeric(apis_df.get("Avg ResTime in sec", 0), errors="coerce").fillna(0).mean()), 3)
        if not apis_df.empty
        else 0
    )

    metrics = [
        ("Total APIs", total_apis),
        ("Total Transactions", total_tx),
        ("Total Error Count", total_error_count),
        ("SLA Pass APIs", sla_pass),
        ("SLA Fail APIs", sla_fail),
        ("Avg API Resp Time Sec", avg_resp),
    ]

    start_row = 3
    for idx, (label, value) in enumerate(metrics):
        row = start_row + idx
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)

    ws["D3"] = "SLA Status"
    ws["E3"] = "Count"
    ws["D4"] = "PASS"
    ws["E4"] = sla_pass
    ws["D5"] = "FAIL"
    ws["E5"] = sla_fail

    pie = PieChart()
    pie.title = "API SLA Pass vs Fail"
    labels = Reference(ws, min_col=4, min_row=4, max_row=5)
    data = Reference(ws, min_col=5, min_row=3, max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 6
    pie.width = 8
    ws.add_chart(pie, "G3")

    top_slow = apis_df.copy()
    top_slow["Avg ResTime in sec"] = pd.to_numeric(top_slow["Avg ResTime in sec"], errors="coerce")
    top_slow = top_slow.sort_values("Avg ResTime in sec", ascending=False).head(10)

    top_start = 12
    ws.cell(row=top_start, column=1, value="Top 10 Slow APIs")
    ws.cell(row=top_start, column=1).font = Font(bold=True, color="1F4E78")
    ws.cell(row=top_start + 1, column=1, value="API")
    ws.cell(row=top_start + 1, column=2, value="Avg Sec")
    for idx, (_, row) in enumerate(top_slow.iterrows(), start=top_start + 2):
        ws.cell(row=idx, column=1, value=f"{row.get('Feature','')}/{row.get('Scenario','')}")
        ws.cell(row=idx, column=2, value=float(row.get("Avg ResTime in sec") or 0))

    bar = BarChart()
    bar.title = "Top 10 Slow APIs"
    bar.y_axis.title = "Avg Sec"
    bar.x_axis.title = "API"
    data = Reference(ws, min_col=2, min_row=top_start + 1, max_row=top_start + 11)
    cats = Reference(ws, min_col=1, min_row=top_start + 2, max_row=top_start + 11)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 9
    bar.width = 18
    ws.add_chart(bar, "G18")

    # Top 10 Error APIs table and chart.
    err_start = 30
    ws.cell(row=err_start, column=1, value="Top 10 Error APIs")
    ws.cell(row=err_start, column=1).font = Font(bold=True, color="9C0006")
    ws.cell(row=err_start + 1, column=1, value="API")
    ws.cell(row=err_start + 1, column=2, value="Error Count")

    top_errors = apis_df.copy()
    top_errors["errorCount"] = pd.to_numeric(top_errors.get("errorCount", 0), errors="coerce").fillna(0)
    top_errors = top_errors[top_errors["errorCount"] > 0].sort_values("errorCount", ascending=False).head(10)

    if top_errors.empty:
        ws.cell(row=err_start + 2, column=1, value="No API errors found")
        ws.cell(row=err_start + 2, column=2, value=0)
    else:
        for idx, (_, row) in enumerate(top_errors.iterrows(), start=err_start + 2):
            ws.cell(row=idx, column=1, value=f"{row.get('Feature','')}/{row.get('Scenario','')}/{row.get('Endpoint','')}")
            ws.cell(row=idx, column=2, value=int(row.get("errorCount", 0)))

        err_bar = BarChart()
        err_bar.title = "Top 10 Error APIs"
        err_bar.y_axis.title = "Error Count"
        err_bar.x_axis.title = "API"
        err_data = Reference(ws, min_col=2, min_row=err_start + 1, max_row=err_start + 1 + len(top_errors))
        err_cats = Reference(ws, min_col=1, min_row=err_start + 2, max_row=err_start + 1 + len(top_errors))
        err_bar.add_data(err_data, titles_from_data=True)
        err_bar.set_categories(err_cats)
        err_bar.height = 8
        err_bar.width = 15
        ws.add_chart(err_bar, "G35")

    style_sheet(ws)



def add_track_comparison_charts(ws):
    """
    Add charts inside Track_Comparison.
    For each uploaded run block, create a small helper table below the comparison table
    showing top 10 tracks by slowest bucket percentage for Avg metric, then chart it.
    """
    if ws.max_row < 3 or ws.max_column < 7:
        return

    source_last_row = ws.max_row
    source_last_col = ws.max_column

    chart_start_row = source_last_row + 4
    ws.cell(row=chart_start_row, column=1, value="Track Comparison Charts")
    ws.cell(row=chart_start_row, column=1).font = Font(size=14, bold=True, color="1F4E78")

    run_block_index = 0
    col = 3

    while col <= source_last_col:
        run_label = ws.cell(row=1, column=col).value
        if not run_label:
            col += 1
            continue

        slow_bucket_col = col + 3
        max_seconds_col = col + 4

        avg_rows = []
        current_track = None

        for row in range(3, source_last_row + 1):
            track_cell = ws.cell(row=row, column=1).value
            metric = ws.cell(row=row, column=2).value

            if track_cell not in (None, ""):
                current_track = track_cell

            if metric == "Avg" and current_track:
                track_normalized = str(current_track).strip().lower()
                if track_normalized == "total" or "select customer" in track_normalized:
                    continue

                slow_pct = ws.cell(row=row, column=slow_bucket_col).value
                max_seconds = ws.cell(row=row, column=max_seconds_col).value

                try:
                    slow_pct = float(slow_pct)
                except Exception:
                    slow_pct = 0.0

                try:
                    max_seconds = float(max_seconds)
                except Exception:
                    max_seconds = 0.0

                avg_rows.append((current_track, slow_pct, max_seconds))

        if not avg_rows:
            col += 6
            continue

        avg_rows = sorted(avg_rows, key=lambda x: (x[1], x[2]), reverse=True)[:10]

        # Create helper table for this run in its own column area.
        table_col = 1 + (run_block_index * 5)
        table_row = chart_start_row + 2

        ws.cell(row=table_row, column=table_col, value=f"{run_label} - Top Slow Bucket %")
        ws.cell(row=table_row, column=table_col).font = Font(bold=True, color="9C0006")

        ws.cell(row=table_row + 1, column=table_col, value="Track")
        ws.cell(row=table_row + 1, column=table_col + 1, value="Slow Bucket %")
        ws.cell(row=table_row + 1, column=table_col + 2, value="Max Seconds")

        for idx, (track, slow_pct, max_seconds) in enumerate(avg_rows, start=table_row + 2):
            ws.cell(row=idx, column=table_col, value=track)
            ws.cell(row=idx, column=table_col + 1, value=slow_pct)
            ws.cell(row=idx, column=table_col + 2, value=max_seconds)

        chart = BarChart()
        chart.title = f"{run_label} - Top Slow Tracks"
        chart.y_axis.title = "Slow Bucket %"
        chart.x_axis.title = "Track"

        data = Reference(
            ws,
            min_col=table_col + 1,
            min_row=table_row + 1,
            max_row=table_row + 11,
        )
        cats = Reference(
            ws,
            min_col=table_col,
            min_row=table_row + 2,
            max_row=table_row + 11,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 14

        chart_anchor_col = get_column_letter(table_col)
        ws.add_chart(chart, f"{chart_anchor_col}{table_row + 14}")

        run_block_index += 1
        col += 6


def write_track_comparison_sheet(wb: Workbook, track_matrix: List[List[Any]]):
    ws = wb.create_sheet("Track_Comparison")
    for row in track_matrix:
        ws.append(row)

    # Merge each run label over its 5 columns, leaving the spacer column unmerged.
    col = 3
    while col <= ws.max_column:
        if ws.cell(row=1, column=col).value:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 4)
        col += 6

    add_track_comparison_charts(ws)
    style_sheet(ws)


def write_excel(frames: Dict[str, pd.DataFrame], output_excel_path: str | Path, track_matrix: List[List[Any]], insights_frames: Dict[str, pd.DataFrame] | None = None, comparison_mode: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    build_insights_sheet(ws, insights_frames if insights_frames is not None else frames)

    # Put Track_Comparison near top after Insights.
    write_track_comparison_sheet(wb, track_matrix)

    sheet_order = ["APIs_Comparison"] if comparison_mode else ["Transactions", "Errors", "APIs", "Comparison"]
    for sheet_name in sheet_order:
        if sheet_name not in frames:
            continue
        ws = wb.create_sheet(sheet_name)
        df = frames[sheet_name]
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append([None if pd.isna(v) else v for v in row.tolist()])
        style_sheet(ws)

    wb.save(output_excel_path)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate Excel report from JMeter statistics.json")
    parser.add_argument("json_files", nargs="+", help="One or more JMeter statistics.json files")
    parser.add_argument("--output", default="JMeter_Report.xlsx", help="Path to output .xlsx file")
    args = parser.parse_args()

    labels = [Path(p).stem for p in args.json_files]

    if len(args.json_files) == 1:
        build_report(args.json_files[0], args.output)
    else:
        build_comparison_report(args.json_files, labels, args.output)
